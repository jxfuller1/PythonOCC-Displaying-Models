import sys
from PyQt5.QtWidgets import QSplashScreen, QApplication
from PyQt5.QtGui import QPixmap


app = QApplication(sys.argv)
splash = QSplashScreen(QPixmap("O:\\put splash img location here.........."))
splash.show()

# shows splashscreen while main program loads

class SplashScreen(QSplashScreen):
    def __init__(self):
        super(QSplashScreen, self).__init__()
        img_path = "O:\\put splash img location here.........."
        self.setWindowFlag(Qt.FramelessWindowHint)
        pixmap = QPixmap(img_path)
        self.setPixmap(pixmap)

    #shows splashscreen while main program loads


from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import numpy as np
import os
import pandas as pd
import subprocess
import time
from OCC.Core.RWStl import rwstl_ReadFile
from OCC.Core.AIS import AIS_Shape
from OCC.Core.BRep import BRep_Builder
from OCC.Core.Graphic3d import Graphic3d_TypeOfShadingModel
from OCC.Core.Quantity import *
from OCC.Core.TopoDS import TopoDS_Face
from OCC.Display.backend import load_backend
load_backend("qt-pyqt5")
from OCC.Display import qtDisplay
import win32gui
from openpyxl import load_workbook


#General outline function activation of program
#--init--
#   --total_buttons
#   --move_pushbuttons
#       --creation of QGraphicsViews for lines
#       --creates buttons/labels
#       --links buttons to onstartclick function
#--onStartClick--
#   --removes child windows of mainwindow minus Qwidgets
#   --recreates center image and new image at button location that was clicked
#   --animates images and calls OnButtonClick when finished
#--onButtonClick--
#   --populates right_layout for GUI
#   --total_buttons
#   --move_pushbuttons
#       --recreation of QGraphicsViews for lines
#       --creates buttons/labels
#       --links buttons to onstartclick function
#       --loads_model function widget over top if center image if model file below 50mb
#--side_bar--
#   --when side bar button clicked
#       --removes child windows of mainwindow minus QWidgets
#       --recreates center image and new image at button location that was clicked
#       --total_buttons
#       --move_pushbuttons
#           --recreation of QGraphicsViews for lines
#           --creates buttons/labels
#           --links buttons to onstartclick function
#           --loads_model function widget over top if center image if model file below 50mb
#--onFindPart--
#   --removes child windows of mainwindow minus QWidgets
#       --recreates center image and new image at button location that was clicked
#       --total_buttons
#       --move_pushbuttons
#           --recreation of QGraphicsViews for lines
#           --creates buttons/labels
#           --links buttons to onstartclick function
#           --loads_model function widget over top if center image if model file below 50mb
#       --populates heirarchy for side bar by calling function that does it
#--open_drawing--
#   --reads drawings folder top open latest drawing
#--open_model--
#   --opens model in program




dir_path = os.path.dirname(os.path.realpath(__file__))
#this is needed for side_bar button background image locations, somehow sets the correct path relative to python interpretor, can't set background image otherwise

#GUI class
class Actions(QDialog):
    
    def __init__(self):
        super(Actions, self).__init__()
        self.initUI()
        self.path_tree = []

        #path_tree variable here as basically a global within Actions for keeping tracking of side_bar pathway as users click down the heirarchy

    # this function for reading excel to get number of parts available
    def OJ_QoH(self, part_number):

        OJ_data = "N/A"
        QoH_data = "N/A"

        #this is for loading a text file that contains items i want to be listed with "NA" in the GUI
        notissued_items = open("O:\\put splash txt file location here..........","r")
        list_notissued_items = notissued_items.readlines()

        if any(part_number in x for x in list_notissued_items):
            OJ_data = "N/A"
            QoH_data = "N/A"

        # this reads text file to reference parts that i want to be listed with an NA

        else:
            if self.dataframe1_bool == True:
                #if excel file exists at location, refer to init function where it sets this variable
                total_QoH = []

                #read excel for getting number of parts in inventory
                for index, row in self.QoH.iloc[:, 0].iteritems():
                    if part_number in str(row):
                        number = self.QoH.iloc[index, 6]
                        total_QoH.append(int(number))

                sum_test = all([isinstance(item, int) for item in total_QoH])
                #a check to make sure all items in the list are integers before calling sum function

                if sum_test == True:
                    total_sum = sum(total_QoH)
                    QoH_data = str(total_sum)
                else:
                    QoH_data = "Error"
                    #if not all integers for summation return Error for variable, this means its not reading correct columns in the excel

                #returns total Quantity On Hand

            if self.dataframe2_bool == True:
                # if excel file exists at location, refer to init function where it sets this variable
                openjobs_number = 0
                for index, row in self.jobsopen.iloc[:, 15].iteritems():
                    if part_number in str(row):
                        openjobs_number += 1

                OJ_data = str(openjobs_number)

                #returns total Open Jobs

        #returns total in inventory / on hand from an excel sheet
        return (OJ_data, QoH_data)


    def total_buttons(self, AKnumber):
        #this function purely for finding how many of certain parts in a Excel so it can return the number to function that gets angle for distributing buttons around center in GUI

        QApplication.setOverrideCursor(Qt.WaitCursor)
        #set cursor to loading symbol, this doesn't really seem to be working that well with the OCCT qt5 model display widget though
        time.sleep(.5)

        self.previous_part = AKnumber
        #this variable for passing previous part number to onstartClick method and other functions that need this later

        #this code purely for my needs, you may not need this, you'll need to adjust this
        part = self.previous_part[0:10]

        #set excel path folder location here, i had more code for determining locations, but that was purely for me
        pathfai = "O:\\enter excel path location here....."

        self.list_ak = []

        if os.path.exists(pathfai):
            read_dir = os.listdir(pathfai)
            file_ext = ".xlsx"

            #read for filename in directory and return true if excel exists with that name
            test_part = any(part in x for x in read_dir)
            test_ext = any(file_ext in x for x in read_dir)
            if test_part == True:
                if test_ext == True:

                    file = [s for s in read_dir if ".xlsx" in s]
            #check if dir exists, read directory and put all matching file extension into list

                    k = 0
                    total_file = len(file)
                    while k < total_file:
                        if file[k][0:2] != "AK":
                            if file[k][0:2] != "PK":
                                if file[k][0:2] != "HK":
                                    del file[k]
                                    total_file -= 1
                                    k -= 1
                        k +=1
            # code above removes hidden or open excel files from file extension list
            
                    full_path = pathfai + "\\" + file[-1]

                    #load that excel file found in the directory and start reading the excel file to find parts listed in it
                    book = load_workbook(full_path, data_only=True)

                    totalsheetsfair1 = []
                    listofsheetnames = book.sheetnames
                    for sheet in book.worksheets:
                        a = str(sheet)
                        if "FAIR 1" in a:
                            totalsheetsfair1.append(sheet)
                        if "Fair 1" in a:
                            totalsheetsfair1.append(sheet)
          
        #This code separates fair sheets on their own lists for iteration
        
                    b = len(totalsheetsfair1)

                    sheetListfair1 = book.sheetnames[0:b]

        #this code gets the proper sheet names list for iteration

                    #check for parts with certain names in the column
                    part_check = ["AK", "HK", "PK"]
                    for sheet in sheetListfair1:
                        ws = book[sheet]
                        for col in ws.iter_cols(max_col =1):
                            for cell in col:
                                if cell.value != None:
                                    if part not in str(cell.value):
                                        cell_string = str(cell.value)
                                        cell_string.strip(" ")
                                        if cell_string[0:2] in part_check:
                                            #append list_ak for use later for populating name of buttons dynamically in gui
                                            self.list_ak.append(cell.value[0:10])

        #code above for reading excel could be done easier through a panda dataframe and also would be faster

        #return total list of parts for populating total buttons to be displayed
        return len(self.list_ak)

       
        
    def get_coordinates(self, a, b, r, angle):
        #this function/equations for calculating distribution of buttons in a circle in GUI based on number of AKs/HK/PK read in excel in total_buttons function
        x = a + r * np.cos(angle*np.pi/180)
        y = b + r * np.sin(angle*np.pi/180)
        return int(x), int(y)

    def move_pushButtons(self, a, b, r, angle):
        #this functions creates all the buttons at given locations (based on how many buttons there are) , creates QGraphicsView for drawing the lines, reads excel to get part names

        self.view = QGraphicsView(self)
        self.view.setScene(QGraphicsScene(self))

        self.view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.view.setFrameShape(QFrame.NoFrame)
        self.view.setRenderHints(QPainter.Antialiasing)
        #must remove scrollbar and do noframe or GUI looks bad

        self.view.setSceneRect(QRectF(0, 0, 1260,910))
        self.view.setGeometry(0, 0, 1260,910)
        self.view.setStyleSheet("background: lightsteelblue")
        self.view.show()

        OJ, QoH = self.OJ_QoH(self.previous_part)
        #uses variable previous_part from total_buttons function

        #add labels for part displayed in center of GUI
        OJQOH_string = "OJ = " + OJ + "     QoH = " + QoH
        self.label_OJQOH = QLabel(OJQOH_string, self)
        myfont = QFont()
        myfont.setBold(True)
        self.label_OJQOH.setFont(myfont)
        self.label_OJQOH.setStyleSheet("font-size: 8pt;")
        self.label_OJQOH.move(int(self.frameGeometry().width()/2 - 425/2 - 62), int(self.frameGeometry().height()/2-325/2 - 18))
        self.label_OJQOH.show()
        #raise labels to the top so they don't get hidden
        self.label_OJQOH.raise_()

        if self.total_ak_level < 28:
            interval = angle
        
        if 27 < self.total_ak_level < 51:
            angle = 360 / 28
            angle1 = 360 / 23
            interval = angle
            interval1 = angle1

        if self.total_ak_level > 50:
            angle = 360 / 28
            angle1 = 360 / 23
            interval = angle
            interval1 = angle1
            width = 75
            height = 40
            height1 = 40   
            height2 = 40
            height3 = self.frameGeometry().height() - 185

        #need the above logic because if too many buttons need to be created they need to be set a certain way for the GUI


        for i in range(0, self.total_ak_level):
      
            if i < 28:
                x, y = self.get_coordinates(a, b, r, angle)
            if 27 < i < 51:
                r = 325
                x, y = self.get_coordinates(a, b, r, angle1)
            if 50 < i < 64:
                x = width
                y = height
            if 63 < i < 77:
                x = self.frameGeometry().width() - 167
                y = height1
            if 76 < i < 79:
                x = self.frameGeometry().width() - 267
                y = height2
            if i > 78:
                x = self.frameGeometry().width() - 267
                y = height3

            # need the above logic because if too many buttons need to be created they needs to be set a certain way for the GUI

            z = x - 62

            self.button = QPushButton(self)
            self.button.setText(self.list_ak[i])
            self.button.setGeometry(z, y-20, 75, 22)
  
            self.button.setFlat(True)
            self.button.setAutoFillBackground(True)
            self.button.setStyleSheet("QPushButton {background-color: white; border-radius : 10; border : 2px solid black; } QPushButton::hover {background-color : lightgrey;}")

            #this path location is for images to be displayed on the GUI for every part
            path = "O:\\setpathway for images" + self.list_ak[i] +".jpg"
            
            if os.path.isfile(path) == True:
                self.button.setToolTip('<br><img src="%s" width="100" height="65">' % ("O:\\setpathway for images" + self.list_ak[i] +".jpg"))
            else:
                self.button.setToolTip('<br><img src="%s" width="100" height="65">' % ("O:\\setpathway for if no image found......noimage.jpg"))

            #set every button to be able to activate onstartclick function
            self.button.clicked.connect(self.onstartClick)
            self.button.show()

            # read excel that was exported to get part names, this could be done faster using panda dataframes
            name_value = " "
            for sheet in self.book_names.sheetnames:
                ws = self.book_names[sheet]
                for row in ws.iter_rows(max_col =1):
                    for cell in row:
                        if cell.value == self.list_ak[i]:
                            row_needed = cell.row  #could probably delete this line not used anywhere
                            name_value = ws.cell(row=cell.row, column = 2).value

            # set name and location just below each button in GUI
            self.label_name = QLabel(self)
            self.label_name.setText(name_value)

            self.label_name.move(z, y+2)
            self.label_name.setStyleSheet("font-size: 6pt;")

            self.label_name.resize(80, 45)
            self.label_name.setWordWrap(True)
            self.label_name.setAlignment(Qt.AlignTop)
            self.label_name.show()

            # creates the lines from button to center, adds to scene to make viewable
            start_x = self.button.x() + self.button.frameGeometry().width()/2
            start_y = self.button.y() + self.button.frameGeometry().height()/2

            start = QPointF(start_x, start_y)
            end = QPointF(self.frameGeometry().width()/2 - 62, self.frameGeometry().height()/2)
            self.view.scene().addItem(QGraphicsLineItem(QLineF(start, end)))
            self.view.lower()

            # finds OJ/QOH for each button item in GUI, this is for inventory amounts
            OJ, QoH = self.OJ_QoH(self.list_ak[i])

            OJ_string = "OJ = " + OJ
            self.label_oj = QLabel(OJ_string, self)
            myfont = QFont()
            myfont.setBold(True)
            self.label_oj.setFont(myfont)
            self.label_oj.setStyleSheet("font-size: 6pt;")
            self.label_oj.move(z-5, y-30)
            self.label_oj.show()

            QoH_string = "QoH = " + QoH
            self.label_QoH = QLabel(QoH_string, self)
            self.label_QoH.setFont(myfont)
            self.label_QoH.setStyleSheet("font-size: 6pt;")
            self.label_QoH.move(z+35, y - 30)
            self.label_QoH.show()

            #get OJ/QOH quantities in inventory from reading excels

            if i < 28:  
                angle += interval
            if 27 < i < 51:
                angle1 += interval1
            if 50 < i < 64:
                height += 65
            if 63 < i < 77:
                height1 += 65
            if 76 < i < 79:
                height2 += 65
            if i > 78:
                height3 += 65
            # need the above logic because if too many buttons need to be created they needs to be set a certain way for the GUI
        QApplication.restoreOverrideCursor()
        #end load cursor icon, this functionality doens't work so well with OCCT display widget

        
    def initUI(self):
        #this function starts when program starts, creates GUI and reads all neccessary excels up front

        self.dataframe1_bool = False
        if os.path.isfile("S:\\excel path location needed here for finding how many in inventory, .....xls"):
            self.dataframe1 = pd.ExcelFile("S:\\excel path location needed here for finding how many in inventory.......xls")
            self.dataframe1_bool = True
            self.names1 = self.dataframe1.sheet_names
            self.QoH = self.dataframe1.parse(self.names1[0])

        #code puts excel into a dataframe if exists  for lookup in OJ_QoH function, this excel gets QoH items

        self.dataframe2_bool = False
        if os.path.isfile("S:\\excel path location needed here for finding how many open jobs for the part.......xls"):
            self.dataframe2 = pd.ExcelFile("S:\\excel path location needed here for finding how many open jobs for the part.......xls")
            self.dataframe2_bool = True
            self.names2 = self.dataframe2.sheet_names
            self.jobsopen = self.dataframe2.parse(self.names2[0])

        #code puts excel into a dataframe if exists for lookup in OJ_QoH function, this excel gets QoH items (quanitty on hand)

        part_names = "O:\\this is for the location of the excel that contains the names of all the parts to display in GUI.......xlsx"
        self.book_names = load_workbook(part_names, data_only=True)

        #reads excel containting part names up front at start of program so it doesn't have to be done later

        self.setGeometry(400,75,1260,910)
        self.setFixedSize(self.size())
        self.setWindowFlag(Qt.WindowMinimizeButtonHint, True)
        self.setWindowTitle("Interactive Plane Parts Diagram")
        #create window and set size
        
        width_center = self.frameGeometry().width() / 2
        height_center = self.frameGeometry().height() / 2
        
        width = self.frameGeometry().width() 
        height = self.frameGeometry().height()

        #enter what you want the starting part in the GUI to be here when program launches
        self.total_ak_level = self.total_buttons("********")
        
        distribution_angle = 360 / self.total_ak_level
        self.move_pushButtons(width_center-50, height_center-20, 415, distribution_angle)

        self.label = QLabel(self)
        pixmap = QPixmap("O:\\enter starting part image file location.jpg")
        smaller_pixmap = pixmap.scaled(425, 325)
        self.label.setPixmap(smaller_pixmap)
        self.label.resize(425, 325)

        width_label = self.label.frameGeometry().width()
        height_label = self.label.frameGeometry().height()

        self.label.move(int((width/2) - (width_label/2 + 62)), int((height/2) - (height_label/2)))
        #set center image


        # create max/min center image/model buttons so user can make bigger
        self.maximize = QPushButton("+", self)
        self.maximize.setStyleSheet("background-color: lightgrey; font-size: 8pt;")
        self.maximize.setFixedSize(QSize(15, 15))
        self.maximize.move(int((self.frameGeometry().width() / 2) - (425 / 2 + 62)) + 410, int((self.frameGeometry().height() / 2) - (325 / 2)) - 18)
        self.maximize.clicked.connect(self.makemax)

        self.minimize = QPushButton("-", self)
        self.minimize.setStyleSheet("background-color: lightgrey; font-size: 8pt;")
        self.minimize.setFixedSize(QSize(15, 15))
        self.minimize.move(int((self.frameGeometry().width() / 2) - (425 / 2 + 62)) + 392, int((self.frameGeometry().height() / 2) - (325 / 2)) - 18)
        self.minimize.clicked.connect(self.makemin)

        #set name of starting part displayed in center of GUI when program launches
        self.label2 = QLabel("Name Here", self)
        myfont = QFont()
        myfont.setBold(True)
        self.label2.setFont(myfont)
        self.label2.setStyleSheet("font-size: 9pt;")
        self.label2.move(self.label.x() + 70, self.label.y() -45)

        #set center label
                    
           
        p = self.palette()
        p.setColor(self.backgroundRole(), QColor("lightsteelblue"))
        self.setPalette(p)
        #set background color of main window

        self.layout = QVBoxLayout()
        self.layout.setAlignment(Qt.AlignTop)
        self.layout.setSpacing(15)


        self.widget = QWidget(self)
        self.widget.setAttribute(Qt.WA_StyledBackground, True)
        self.widget.setStyleSheet('background-color: steelblue;')
        self.widget.setGeometry(width-145, 0, 145, height-40)

        self.widget.setLayout(self.layout)
        #set side_bar layout


        self.layout_bottom = QHBoxLayout()
        self.layout_bottom.setAlignment(Qt.AlignCenter)
        
        self.input_part = QLineEdit(self)
        self.input_part.setPlaceholderText("Enter part number")
        self.input_part.setStyleSheet('background-color: white;')

        self.input_button = QPushButton("Go", self)
        self.input_button.setStyleSheet('background-color: lightgrey;')
        self.input_button.setFixedSize(QSize(25, 20))
        
        self.input_button.clicked.connect(self.onfindPart)
    
        self.layout_bottom.addWidget(self.input_part)
        self.layout_bottom.addWidget(self.input_button)

        self.widget_bottom = QWidget(self)
        self.widget_bottom.setAttribute(Qt.WA_StyledBackground, True)
        self.widget_bottom.setStyleSheet('background-color: steelblue;')
        self.widget_bottom.setGeometry(width-145, height-40, 145, 40)
        self.widget_bottom.setLayout(self.layout_bottom)
        # center bottom right layout and input for part number and connect to function
        

        self.widget_drawing = QWidget(self)
        self.widget_drawing.setAttribute(Qt.WA_StyledBackground, True)
        self.widget_drawing.setStyleSheet('background-color: steelblue;')
        self.widget_drawing.setGeometry(0, height-40, 145, 40)

        self.drawing_layout = QVBoxLayout()
        self.drawing_layout.setAlignment(Qt.AlignCenter)
        
        self.drawing_button = QPushButton("Open Latest Drawing", self)
        self.drawing_button.setFixedSize(QSize(125, 20))
        self.drawing_button.setStyleSheet('background-color: lightgrey;')
        self.drawing_button.clicked.connect(self.open_drawing)

        self.drawing_layout.addWidget(self.drawing_button)
        self.widget_drawing.setLayout(self.drawing_layout)

        #set bottom left open drawings layout and button and connect to function


        self.widget_model = QWidget(self)
        self.widget_model.setAttribute(Qt.WA_StyledBackground, True)
        self.widget_model.setStyleSheet('background-color: steelblue;')
        self.widget_model.setGeometry(145, height - 40, 145, 40)

        self.model_layout = QVBoxLayout()
        self.model_layout.setAlignment(Qt.AlignLeft)

        self.model_button = QPushButton("Open Model", self)
        self.model_button.setFixedSize(QSize(125, 20))
        self.model_button.setStyleSheet('background-color: lightgrey;')
        self.model_button.clicked.connect(self.open_model)

        self.model_layout.addWidget(self.model_button)
        self.widget_model.setLayout(self.model_layout)

        #set bottom left open model button and connect to function


        self.widget_legend = QWidget(self)
        self.widget_legend.setAttribute(Qt.WA_StyledBackground, True)
        self.widget_legend.setStyleSheet('background-color: steelblue;')
        self.widget_legend.setGeometry(width-520, height - 40, 375, 40)

        self.legend_layout = QVBoxLayout()
        self.legend_layout.setAlignment(Qt.AlignCenter)

        string_legend = "Legend: \nOJ = Open Jobs      QoH = Quantity On Hand       N/A = Not Applicable"

        self.qoh_label = QLabel(string_legend)
        self.qoh_label.setFont(myfont)
        self.qoh_label.setStyleSheet("font-size: 7pt;")

        self.legend_layout.addWidget(self.qoh_label)
        self.widget_legend.setLayout(self.legend_layout)

        #creates bottom right legend layout on GUI

        self.NHAs = load_workbook("O:\\excel location for heirarchy of parts list.xlsx", data_only=True)

        #loads excel on opening of program so it doesn't have to be done later



    def onstartClick(self, value):
        #function starts when part button clicked in GUI
            
        self.view.scene().clear()
        #clear graphic lines on click of part button in GUI
        
        sender = self.sender()
        #gets the widget/button that was clicked on
        
        m = int(sender.x() + (sender.frameGeometry().width() / 2)) 
        z = int(sender.y() + (sender.frameGeometry().height() / 2))

        #button location used later for setting image there for animation in GUI
                      
        for widget in self.children():
            #reads all children widgets (but not QWidgets) on the main window GUI and removes them with isinstance statements

            if isinstance(widget, qtDisplay.qtViewer3d):
                widget.deleteLater()
                #deletes center model view widget if the widget exists

            if isinstance(widget, QGraphicsView):
                widget.deleteLater()
                #deletes qgraphicsview , will be re-created later, this is only used for creating lines

            if isinstance(widget, QLabel):
                widget_test = len(widget.text())
                if widget_test != 0:
                    widget.deleteLater()
                    #deletes all labels, part names, OJ, QOH from main window
                
            if isinstance(widget, QPushButton):
                if sender.text() not in widget.text():
                    widget.deleteLater()
                    #delete all buttons for the one that wasn't clicked

                if sender.text() in widget.text():
                    #for the one that was clicked do below stuff
                    widget.deleteLater()
                

                    width = self.frameGeometry().width() 
                    height = self.frameGeometry().height()
                    
                    self.label1 = QLabel(self)

                    path = "O:\\location of image for part here...." + self.previous_part + ".jpg"
                    
                    if os.path.isfile(path) == False:
                        pixmap1 = QPixmap("O:\\location of image for part here if image not found....noimage.jpg")
                    else:
                        pixmap1 = QPixmap("O:\\location of image for part here...." + self.previous_part + ".jpg")
                        
                    #smaller_pixmap1 = pixmap1.scaled(325, 225, Qt.KeepAspectRatio)
                    smaller_pixmap1 = pixmap1.scaled(425, 325)               
                    self.label1.setPixmap(smaller_pixmap1)
                            
                    self.label1.resize(425, 325)
  
                    width_label1 = self.label1.frameGeometry().width()
                    height_label1 = self.label1.frameGeometry().height()

                    self.label1.move(int((width/2) - (width_label1/2 + 62)), int((height/2) - (height_label1/2)))

                    #above code creates new image of current part in center to move to side bar for animation


                    path = "O:\\location of image for part here...." + sender.text() + ".jpg"
                    if os.path.isfile(path) == False:
                        pixmap = QPixmap("O:\\location of image for part here if image not found....noimage.jpg")
                    else:
                        pixmap = QPixmap("O:\\location of image for part here...." + sender.text() + ".jpg")

                    smaller_pixmap = pixmap.scaled(425, 325)
                    self.label.setPixmap(smaller_pixmap)
                    self.label.resize(425, 325)

                    self.label.move(abs(int(m - (self.label.frameGeometry().width() / 2))) , abs(int(z - (self.label.frameGeometry().height() / 2))))

                    #above code creates new image of new part at button location that was clicked to move to center for animation

                    self.label1.show()
                    self.label.show()

                    self.widget.raise_()

                    #raises sidebar widget/layout to top so image appears to go under it when animation starts

                                       
                    
        self.anim = QParallelAnimationGroup()

        #set animation locations to move to in "first" and "second"
        first = QPropertyAnimation(self.label, b"pos")
        second = QPropertyAnimation(self.label1, b"pos")


        m = int(self.label.x() + (self.label.frameGeometry().width() / 2)) 
        z = int(self.label.y() + (self.label.frameGeometry().height() / 2))


        move_button_x = abs(int((self.frameGeometry().width() / 2 - 62) - (self.label.frameGeometry().width() / 2)))
        move_button_y = abs(int((self.frameGeometry().height() / 2) - (self.label.frameGeometry().height() / 2)))

        #getting the x, y value difference to move image to center from where the button is at

        first.setEndValue(QPoint(move_button_x, move_button_y))                          
        first.setDuration(700)

        #moves new image to center of screen from whichever button was clicked

        second.setEndValue(QPoint(self.frameGeometry().width() - 115, -100))
        second.setDuration(700)

        #moves center image to side bar

        self.anim.addAnimation(first)
        self.anim.addAnimation(second)

        self.anim.start()

        self.anim.finished.connect(self.onButtonClick)
        #start next function for repopulating GUI based on what button was clicked

        self.test_sender = sender.text()
        #for pass through to onbuttonclick function in order to repopulate center of GUI


        
    def onButtonClick(self):
        #activates after animation finished in onStartClick

        self.label1.deleteLater()
        #deletes image that was animated to the top right corner for animation

        rightside_layout = self.right_layout(self.previous_part)
        #calls right_layout to add to side_bar path

        #note: self.text_sender from above (it's based on the button that was clicked surround part in center of GUI
        model_path = "O:\\model file location for part........." + self.test_sender + ".stl"
        if os.path.isfile(model_path) == True:
            file_size = os.path.getsize(model_path)

            #this is set so that if model size is above 50Mb it won't load it in program, anything bigger and it starts to take a bit to load
            if file_size < 50000000:
                #call model function to load model
                self.load_model(self.test_sender)
                #lower and raise the canvas for the model to remove issues with it getting hidden
                self.canvas.lower()
                self.canvas.raise_()

                #if models loads
                self.model_loaded_label = QLabel("Interactive Model Loaded!", self)
                self.model_loaded_label.move(int(self.frameGeometry().width() / 2 - 425 / 2 - 62), int(self.frameGeometry().height() / 2 + 325 / 2 + 8))
                self.model_loaded_label.show()

        #set min/max buttons to be able to make window larger in center , especially when viewing a model
        self.maximize = QPushButton("+", self)
        self.maximize.setStyleSheet("background-color: lightgrey; font-size: 8pt;")
        self.maximize.setFixedSize(QSize(15, 15))
        self.maximize.move(int((self.frameGeometry().width() / 2) - (425 / 2 + 62)) + 410, int((self.frameGeometry().height() / 2) - (325 / 2)) - 18)
        self.maximize.clicked.connect(self.makemax)
        self.maximize.show()

        self.minimize = QPushButton("-", self)
        self.minimize.setStyleSheet("background-color: lightgrey; font-size: 8pt;")
        self.minimize.setFixedSize(QSize(15, 15))
        self.minimize.move(int((self.frameGeometry().width() / 2) - (425 / 2 + 62)) + 392, int((self.frameGeometry().height() / 2) - (325 / 2)) - 18)
        self.minimize.clicked.connect(self.makemin)
        self.minimize.show()

        #re-add min / max buttons

        width_center = self.frameGeometry().width() / 2
        height_center = self.frameGeometry().height() / 2

        self.total_ak_level = self.total_buttons(self.test_sender)
        #calls total buttons to get how many buttons to population


        #if parts found with excel
        if self.total_ak_level != 0:  
            distribution_angle = 360 / self.total_ak_level
            self.move_pushButtons(width_center-50, height_center-20, 415, distribution_angle)
            #calls move puthbuttons to populate buttons and calculate where they go in GUI

        #if parts not found within excel set defaul graphics scene without buttons
        if self.total_ak_level == 0:
            QApplication.restoreOverrideCursor()
                
            self.view = QGraphicsView(self)
            self.view.setScene(QGraphicsScene(self))

            self.view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self.view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self.view.setFrameShape(QFrame.NoFrame)
            self.view.setRenderHints(QPainter.Antialiasing)

            self.view.setSceneRect(QRectF(0, 0, 1260,910))
            self.view.setGeometry(0, 0, 1260,910)
            self.view.setStyleSheet("background: lightsteelblue")
            self.view.lower()
            self.view.show()
            #if part has no AK/PK/HKs move_pushbuttons function is not called ,therefore QGrahpicsView needs to be re-added (QGraphicsView is added in move_pushbuttons)

            OJ, QoH = self.OJ_QoH(self.test_sender)

            OJQOH_string = "OJ = " + OJ + "     QoH = " + QoH
            self.label_OJQOH = QLabel(OJQOH_string, self)
            myfont = QFont()
            myfont.setBold(True)
            self.label_OJQOH.setFont(myfont)
            self.label_OJQOH.setStyleSheet("font-size: 8pt;")
            self.label_OJQOH.move(int(self.frameGeometry().width() / 2 - 425 / 2 - 62), int(self.frameGeometry().height() / 2 - 325 / 2 - 18))
            self.label_OJQOH.show()
            self.label_OJQOH.raise_()
            #calls OJ_QOH functions to get QOH/OJ numbers, adds them to a label for center item in GUI


        # reads excel that contains names for parts and get name for center part of gui
        name_value = " "
        for sheet in self.book_names.sheetnames:
            ws = self.book_names[sheet]
            for row in ws.iter_rows(max_col =1):
                for cell in row:
                    if cell.value == self.test_sender:
                        row_needed = cell.row
                        name_value = ws.cell(row=cell.row, column = 2).value

        label2_name = self.test_sender + "  -  " + name_value      
        self.label2 = QLabel(label2_name, self)
        myfont = QFont()
        myfont.setBold(True)
        self.label2.setFont(myfont)
        self.label2.setStyleSheet("font-size: 9pt;")
        self.label2.move(self.label.x() + 70, self.label.y() -45)
        self.label2.show()


    def load_model(self, partnumber):
        #creates OCCT display widget and reads STL model to it
        self.canvas = qtDisplay.qtViewer3d(self)
        self.canvas.resize(425, 325)

        #values to set size of canvas widget for model
        width_label = self.canvas.frameGeometry().width()
        height_label = self.canvas.frameGeometry().height()
        width = self.frameGeometry().width()
        height = self.frameGeometry().height()

        self.canvas.move(int((width / 2) - (width_label / 2 + 62)), int((height / 2) - (height_label / 2)))
        #calculating center spot to move widget to for GUI

        #set canvas display and color
        self.display = self.canvas._display
        self.display.View.SetBgGradientColors(Quantity_Color(Quantity_NameOfColor(Quantity_NOC_LIGHTBLUE)),
                                              Quantity_Color(Quantity_NameOfColor(Quantity_NOC_GRAY0)), 2, True, )

        #model file path here
        model_file = "O:\\model file path" + partnumber + ".stl"

        #read stl file using faster method documented in SWIG file for pythonocc
        stl_shp = rwstl_ReadFile(model_file)

        #this is all needed in order to build the model properly so it can be displayed in the widget display
        aShape = TopoDS_Face()
        aBuilder = BRep_Builder()
        aBuilder.MakeFace(aShape, stl_shp)
        aShapePrs = AIS_Shape(aShape)
        aShapePrs.Attributes().SetupOwnShadingAspect()
        aShapePrs.SetColor(Quantity_Color(Quantity_NameOfColor(Quantity_NOC_GRAY49)))
        aFillAspect = aShapePrs.Attributes().ShadingAspect().Aspect()
        aFillAspect.SetShadingModel(Graphic3d_TypeOfShadingModel.Graphic3d_TOSM_FACET)

        #show canvas in pyqt5 GUI window
        self.canvas.show()
        self.canvas.InitDriver()
        self.display.Context.Display(aShapePrs, False)
        self.display.View.FitAll()
        #raise to make sure it's on top
        self.canvas.raise_()


        #creates graphics window in center, note the last 4 lines show()/InitDriver() and Display/DIsplay.view have to be called in that order
        #or very 1st model loaded will not scale correctly, but only the 1st one is messed up... i don't know why



    def right_layout(self, number):
        #creates button on side_bar for hierarchy
        self.path_tree.append(number)
        #populate list for side_bar heirarchy for button/number that was clicked

        self.button = QPushButton(number, self)

        #file location for images to be used for right side bar layout
        location = "O:\\image location" + number + ".jpg"

        if os.path.isfile(location) == False:
            location = "O:\\file location if no image found.....noimage.jpg"

                
        path_image = os.path.join(dir_path, location).replace("\\", "/")
        #this is needed in order to set a image pic for the button
        qss = 'border-image: url({})'.format(path_image)
        #this has to be used just like this in order to set a image pic for button for right sie layout
        self.button.setStyleSheet(qss)
        self.button.setCursor(QCursor(Qt.PointingHandCursor))
        #this sets a hand cursor when mouse cursor over button, but this doens't seem to work using the OCCT display widget module

        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(15)
        self.button.setGraphicsEffect(shadow)

        self.button.setFixedSize(125, 100)
        self.layout.insertWidget(0, self.button)

        self.button.clicked.connect(self.side_bar)
        #creates button, sets image/effect, connects button to side_bar function



    def side_bar(self):
        #this fucntion activates when button side bar in GUI is clicked
        sender = self.sender()

        self.view.scene().clear()
        #clear lines in QgraphicsView

        #get index location of the button clicked in  the right side layout of the GUI
        index_loc = self.path_tree.index(sender.text())
        total_path_tree = len(self.path_tree)

        #deletes the right side layout buttons for the ones above the one clicked
        while index_loc < total_path_tree:  
            mywidget = self.layout.itemAt(0).widget()
            self.layout.removeWidget(mywidget)
            mywidget.deleteLater()
            
            del self.path_tree[total_path_tree-1]
      
            total_path_tree -=1
        #this while loop removes the side bar buttons in GUI up to the point to whichever button was clicked

        width_center = self.frameGeometry().width() / 2
        height_center = self.frameGeometry().height() / 2

        #remove all child windows directly on mainwindow for repopulation, doesn't remove the QWidgets
        for widget in self.children():
            if isinstance(widget, qtDisplay.qtViewer3d):
                widget.deleteLater()
            if isinstance(widget, QGraphicsView):
                widget.deleteLater()          
            if isinstance(widget, QPushButton):              
                widget.deleteLater()
            if isinstance(widget, QLabel):
                widget.deleteLater()


        #get number of buttons to population
        self.total_ak_level = self.total_buttons(sender.text())

        if self.total_ak_level != 0:
            distribution_angle = 360 / self.total_ak_level
            self.move_pushButtons(width_center-50, height_center-20, 415, distribution_angle)
            #populates buttons at given locations from function

        #if no parts found in excel set window up without populating any buttons
        if self.total_ak_level == 0:
            QApplication.restoreOverrideCursor()
                
            self.view = QGraphicsView(self)
            self.view.setScene(QGraphicsScene(self))

            self.view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self.view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self.view.setFrameShape(QFrame.NoFrame)
            self.view.setRenderHints(QPainter.Antialiasing)

            #self.view.setSceneRect(QRectF(self.view.viewport().rect()))
            self.view.setSceneRect(QRectF(0, 0, 1260,910))
            self.view.setGeometry(0, 0, 1260,910)
            self.view.setStyleSheet("background: lightsteelblue")
            self.view.lower()
            self.view.show()

            #if no buttons to populate, remake QgraphicsView, it's remade if move_pushbuttons already if that is called above

            #call OJ_QOH function to find OJ and QOH for the side_bar item clicked and populate label of it to center
            OJ, QoH = self.OJ_QoH(sender.text())
            OJQOH_string = "OJ = " + OJ + "     QoH = " + QoH
            self.label_OJQOH = QLabel(OJQOH_string, self)
            myfont = QFont()
            myfont.setBold(True)
            self.label_OJQOH.setFont(myfont)
            self.label_OJQOH.setStyleSheet("font-size: 8pt;")
            self.label_OJQOH.move(int(self.frameGeometry().width() / 2 - 425 / 2 - 62), int(self.frameGeometry().height() / 2 - 325 / 2 - 18))
            self.label_OJQOH.show()
            self.label_OJQOH.raise_()


        self.label = QLabel(self)
        #image label location for center of GUI
        path = "O:\\image file location" + sender.text() + ".jpg"

        if os.path.isfile(path) == False:
            pixmap = QPixmap("O:\\if no image found default use location\\noimage.jpg")
        else:
            pixmap = QPixmap("O:\\image file location" + sender.text() + ".jpg")

        #scale image to label size
        smaller_pixmap = pixmap.scaled(425, 325)
        self.label.setPixmap(smaller_pixmap)
                    
        self.label.resize(425, 325)

        #set image location in GUI
        width = self.frameGeometry().width()
        height = self.frameGeometry().height()
        width_label = self.label.frameGeometry().width()
        height_label = self.label.frameGeometry().height()

        self.label.move(int((width / 2) - (width_label / 2 + 62)), int((height / 2) - (height_label / 2)))

        self.label.show()

        #Make new center label with image for item that was clicked in side bar of GUI, this center label has to be labeled as self.label for the animation later

        #model locations to load in center of GUI
        model_path = "O:\\model location path" + sender.text() + ".stl"
        if os.path.isfile(model_path) == True:
            file_size = os.path.getsize(model_path)

            #won't load model over 50mb due to taking too long, can be changed
            if file_size < 50000000:
                #load model by called load_model function
                self.load_model(sender.text())
                self.model_loaded_label = QLabel("Interactive Model Loaded!", self)
                self.model_loaded_label.move(int(self.frameGeometry().width() / 2 - 425 / 2 - 62), int(self.frameGeometry().height() / 2 + 325 / 2 + 8))
                self.model_loaded_label.show()

        #display OCCT display widget for model if model file less that 50mb, puts it over top of center label image, center label image needs to remain for animation later


        #set buttons to make model window larger or smaller for easier viewing
        self.maximize = QPushButton("+", self)
        self.maximize.setStyleSheet("background-color: lightgrey; font-size: 8pt;")
        self.maximize.setFixedSize(QSize(15, 15))
        self.maximize.move(int((self.frameGeometry().width() / 2) - (425 / 2 + 62)) + 410, int((self.frameGeometry().height() / 2) - (325 / 2)) - 18)
        self.maximize.clicked.connect(self.makemax)
        self.maximize.show()

        self.minimize = QPushButton("-", self)
        self.minimize.setStyleSheet("background-color: lightgrey; font-size: 8pt;")
        self.minimize.setFixedSize(QSize(15, 15))
        self.minimize.move(int((self.frameGeometry().width() / 2) - (425 / 2 + 62)) + 392, int((self.frameGeometry().height() / 2) - (325 / 2)) - 18)
        self.minimize.clicked.connect(self.makemin)
        self.minimize.show()

        #get name of part number clicked on side bar from excel conformity to put name in center of GUI
        name_value = " "
        for sheet in self.book_names.sheetnames:
            ws = self.book_names[sheet]
            for row in ws.iter_rows(max_col =1):
                for cell in row:
                    if cell.value == sender.text():
                        row_needed = cell.row
                        name_value = ws.cell(row=cell.row, column = 2).value
                        

        label2_name = sender.text() + "  -  " + name_value      
        self.label2 = QLabel(label2_name, self)
        myfont = QFont()
        myfont.setBold(True)
        self.label2.setFont(myfont)
        self.label2.setStyleSheet("font-size: 9pt;")
        self.label2.move(self.label.x() + 70, self.label.y() - 45)
        self.label2.show()


    def onfindPart(self):
        #this is for entering a part number in GUI and wiping GUI center, repopulating, populating side_bar with hierarchy
        self.part_number = self.input_part.text().upper()

        part = self.part_number[0:10]

        #file location for excel for part to read excel to see what it contains
        folderfai = "O:\\base folder path you made not need this"
        pathfai = "O:\\file location here for excel for part to read it"

        #make sure excel exists and that ext is correct
        if os.path.exists(pathfai):
            read_dir = os.listdir(pathfai)
            file_ext = ".xlsx"
            test_part = any(part in x for x in read_dir)
            test_ext = any(file_ext in x for x in read_dir)
            if test_part == True:
                if test_ext == True:
                    total_path_tree = len(self.path_tree)
                    if total_path_tree > 0:
                        while total_path_tree > 0:
                            mywidget = self.layout.itemAt(0).widget()
                            self.layout.removeWidget(mywidget)
                            mywidget.deleteLater()
            
                            del self.path_tree[total_path_tree-1]
      
                            total_path_tree -=1

        #removes side_bar items in GUI for repopulation

                    for widget in self.children():
                        if isinstance(widget, qtDisplay.qtViewer3d):
                            widget.deleteLater()
                        if isinstance(widget, QGraphicsView):
                            widget.deleteLater()
                        if isinstance(widget, QPushButton):
                            widget.deleteLater()
                        if isinstance(widget, QLabel):
                            widget.deleteLater()

        #removes all direct child widgets that are not a Qwidget of mainwindow

                    width_center = self.frameGeometry().width() / 2
                    height_center = self.frameGeometry().height() / 2

       
                    self.total_ak_level = self.total_buttons(self.part_number)
            #calls function to read excel to get how many buttons there are

                    if self.total_ak_level != 0:
                        distribution_angle = 360 / self.total_ak_level
                        self.move_pushButtons(width_center-50, height_center-20, 415, distribution_angle)
                #calls function to populate center buttons/GUI
                    if self.total_ak_level == 0:
                        QApplication.restoreOverrideCursor()
                
                        self.view = QGraphicsView(self)
                        self.view.setScene(QGraphicsScene(self))

                        self.view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                        self.view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                        self.view.setFrameShape(QFrame.NoFrame)
                        self.view.setRenderHints(QPainter.Antialiasing)

                #self.view.setSceneRect(QRectF(self.view.viewport().rect()))
                        self.view.setSceneRect(QRectF(0, 0, 1260,910))
                        self.view.setGeometry(0, 0, 1260, 910)
                        self.view.setStyleSheet("background: lightsteelblue")
                        self.view.lower()
                        self.view.show()
                #if no buttons to populate for GUI, re-create QgraphicsView, Qgraphicsview is populated in the function of move_pushbutton if it's called

                        OJ, QoH = self.OJ_QoH(part)

                        OJQOH_string = "OJ = " + OJ + "     QoH = " + QoH
                        self.label_OJQOH = QLabel(OJQOH_string, self)
                        myfont = QFont()
                        myfont.setBold(True)
                        self.label_OJQOH.setFont(myfont)
                        self.label_OJQOH.setStyleSheet("font-size: 8pt;")
                        self.label_OJQOH.move(int(self.frameGeometry().width() / 2 - 425 / 2 - 62),
                                      int(self.frameGeometry().height() / 2 - 325 / 2 - 18))
                        self.label_OJQOH.show()
                        self.label_OJQOH.raise_()
                #get QOH/OJ numbers and label in center image

                    self.label = QLabel(self)
                    path = "O:\\path to image file for part" + self.part_number + ".jpg"

                    if os.path.isfile(path) == False:
                        pixmap = QPixmap("O:\\path if no image found.....noimage.jpg")
                    else:
                        pixmap = QPixmap("O:\\path to image file" + self.part_number + ".jpg")
            
                    smaller_pixmap = pixmap.scaled(425, 325)
                    self.label.setPixmap(smaller_pixmap)
                    self.label.resize(425, 325)

                    width = self.frameGeometry().width()
                    height = self.frameGeometry().height()
                    width_label = self.label.frameGeometry().width()
                    height_label = self.label.frameGeometry().height()

                    self.label.move(int((width / 2) - (width_label / 2 + 62)), int((height / 2) - (height_label / 2)))
                    self.label.show()

            #creates center image for GUI based on number typed in GUI, note this must be labeled as self.label for the animation to work properly later


                    model_path = "O:\\path to model file" + self.part_number + ".stl"
                    if os.path.isfile(model_path) == True:
                        file_size = os.path.getsize(model_path)
                        if file_size < 50000000:
                            self.load_model(self.part_number)
                            self.canvas.lower()
                            self.canvas.raise_()
                            self.model_loaded_label = QLabel("Interactive Model Loaded!", self)
                            self.model_loaded_label.move(int(self.frameGeometry().width() / 2 - 425 / 2 - 62), int(self.frameGeometry().height() / 2 + 325 / 2 + 8))
                            self.model_loaded_label.show()

            #this lower/raise purely in here because of a bug that happens when you search for a part immediately after opening program,
            # the middle image doesn't update for some reason on first launch, lower/raise fixes it

                    name_value = " "
                    for sheet in self.book_names.sheetnames:
                        ws = self.book_names[sheet]
                        for row in ws.iter_rows(max_col =1):
                            for cell in row:
                                if cell.value == self.part_number:
                                    row_needed = cell.row
                                    name_value = ws.cell(row=cell.row, column = 2).value
                        

                    label2_name = self.part_number + "  -  " + name_value
                    self.label2 = QLabel(label2_name, self)
                    myfont = QFont()
                    myfont.setBold(True)
                    self.label2.setFont(myfont)
                    self.label2.setStyleSheet("font-size: 9pt;")
                    self.label2.move(self.label.x() + 70, self.label.y() - 45)
                    self.label2.show()

            #get name of part trying to find from excel and add it's name to center in GUI

                    self.maximize = QPushButton("+", self)
                    self.maximize.setStyleSheet("background-color: lightgrey; font-size: 8pt;")
                    self.maximize.setFixedSize(QSize(15, 15))
                    self.maximize.move(int((self.frameGeometry().width() / 2) - (425 / 2 + 62)) + 410,int((self.frameGeometry().height() / 2) - (325 / 2)) - 18)
                    self.maximize.clicked.connect(self.makemax)
                    self.maximize.show()

                    self.minimize = QPushButton("-", self)
                    self.minimize.setStyleSheet("background-color: lightgrey; font-size: 8pt;")
                    self.minimize.setFixedSize(QSize(15, 15))
                    self.minimize.move(int((self.frameGeometry().width() / 2) - (425 / 2 + 62)) + 392,int((self.frameGeometry().height() / 2) - (325 / 2)) - 18)
                    self.minimize.clicked.connect(self.makemin)
                    self.minimize.show()

            #re-add max/min buttons

        #code below to find NHA's and populate into side bar in GUI

                    list_ak = []
                    row_needed = False
                    for sheet in self.NHAs.sheetnames:
                        ws = self.NHAs[sheet]
                        for row in ws.iter_rows(max_col=1):
                            for cell in row:
                                if cell.value == self.part_number:
                                    row_needed = cell.row
                        if row_needed != False:
                            for col in ws.iter_cols(min_col=2, min_row=row_needed, max_row=row_needed):
                                for cell in col:
                                    if ws.cell(row=cell.row, column=cell.column).value != None:
                                        if ws.cell(row=cell.row, column=cell.column).value != "Multiple":
                                            list_ak.append(cell.value)

                    list_ak.reverse()
                    total_aks = len(list_ak)
                    k = 0
                    while k < total_aks:
                        self.right_layout(list_ak[k])
                        k +=1


    def open_drawing(self):
        self.drawing = self.previous_part
        self.find_drawing()
        #function activates when open latest drawing button clicked in GUI, activates next function find_drawing

    def find_drawing(self):
        #reads folder to get latest pdf and open it
        path = "S:\\path to pdf folder"

        #read folder and if more than 1 file with that name, use latest in list
        read_dir = os.listdir(path)
        file = [s for s in read_dir if self.drawing in s]

        total_file = len(file)
        if file != 0:
            latest_drawing = file[-1]

        #gets latest drawing by innately using windows sort order when reading dir, last alphabetical one in list is the one to open and should be the latest drawing
       
        pdf_path = path + "\\" + latest_drawing
        
        os.startfile(pdf_path)


    def open_model(self):
        #activates when opem model button is clicked in GUI

        self.model = self.previous_part

        path = "O:\\path to model file" + self.model + ".stl"
        edrawings_path = "C:\\path to program that can open model files"

        model_path = os.path.isfile(path)

        enquiries = win32gui.FindWindow(None, "eDrawings")

        if model_path == True:
            if os.path.isfile(edrawings_path) == True:
                if enquiries == 0:
                    p = subprocess.Popen([edrawings_path, path])
                else:
                    os.startfile(path)

        #gets current part in center image in GUI and attempts to open model in edrawings, if edrawings not found or is already open,
        # will just straight open the file, which defaults to print3d for windows 11.


    def makemax(self):
        #finds the child QLabel on main window that has no text (theres only 1 in center of screen)
        # makes the center image larger, then creates/moves the max button to the top right of it
        for widget in self.children():
            if isinstance(widget, QLabel):
                widget_test = len(widget.text())
                if widget_test == 0:
                    imagepath = "O:\\image path for part listed" + self.previous_part + ".jpg"
                    pixmap = QPixmap(imagepath)
                    x_scaled = self.frameGeometry().width() - 245
                    y_scaled = self.frameGeometry().height() - 140
                    smaller_pixmap = pixmap.scaled(x_scaled, y_scaled)
                    widget.setPixmap(smaller_pixmap)
                    widget.setGeometry(50, 50, self.frameGeometry().width() - 245, self.frameGeometry().height() - 140)
                    widget.raise_()
            if isinstance(widget, qtDisplay.qtViewer3d):
                widget.setGeometry(50, 50, self.frameGeometry().width() - 245, self.frameGeometry().height() - 140)
                widget.raise_()
                #center display widget if it exists at center to be on top just in case
            if isinstance(widget, QPushButton):
                #widget_test = len(widget.text())
                if widget.text() == "+":
                    widget.move(self.frameGeometry().width()-212, 32)
                    widget.raise_()
                if widget.text() == "-":
                    widget.move(self.frameGeometry().width()-230, 32)
                    widget.raise_()
            #raises the min/max buttons to the top so they don't get underneath other buttons


    def makemin(self):
        # finds the child QLabel on main window that has no text (theres only 1 in center of screen)
        # makes the center image smaller, then creates/moves the max button to the top right of it
        width = self.frameGeometry().width()
        height = self.frameGeometry().height()
        for widget in self.children():
            if isinstance(widget, QLabel):
                widget_test = len(widget.text())
                if widget_test == 0:
                    imagepath = "O:\\image path for part" + self.previous_part + ".jpg"
                    pixmap = QPixmap(imagepath)
                    smaller_pixmap = pixmap.scaled(425, 325)
                    widget.setPixmap(smaller_pixmap)
                    widget.setGeometry(int((width/2) - (425/2 + 62)), int((height/2) - (325/2)), 425, 325)
                    widget.raise_()
            # center display widget if it exists at center to be on top just in case
            if isinstance(widget, qtDisplay.qtViewer3d):
                widget.setGeometry(int((width/2) - (425/2 + 62)), int((height/2) - (325/2)), 425, 325)
                widget.raise_()
            if isinstance(widget, QPushButton):
                if widget.text() == "+":
                    widget.move(int((width/2) - (425/2 + 62)) + 410, int((height/2) - (325/2)) - 18)
                    widget.raise_()
                if widget.text() == "-":
                    widget.move(int((width/2) - (425/2 + 62)) + 392, int((height/2) - (325/2)) - 18)
                    widget.raise_()
                # raises the min/max buttons to the top so they don't get underneath other buttons

if __name__ == "__main__":
    app = QApplication(sys.argv)

    #splashscreen on loadup
    splash = SplashScreen()
    splash.show()

    window = Actions()
    window.show()

    #end splash screen after gui loaded
    splash.finish(window)

    sys.exit(app.exec_())


